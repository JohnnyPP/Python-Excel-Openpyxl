# -*- coding: utf-8 -*-
#http://ericgazoni.wordpress.com/2010/04/10/openpyxl-python-xlsx/
#https://bitbucket.org/ericgazoni/openpyxl/overview
#http://pythonhosted.org/openpyxl/index.html

from openpyxl.reader.excel import load_workbook

filename='130806 4M ADCSI10us Average100 SI1000us slow.xlsx'
#130806 4M ADCSI10us Average100 SI10000us fast.xlsx
#130806 4M ADCSI10us Average100 SI10000us medium.xlsx
#130806 4M ADCSI10us Average100 SI10000us slow.xlsx
#130806 4M ADCSI10us Average100 SI1000us fast.xlsx
#130806 4M ADCSI10us Average100 SI1000us slow.xlsx
#130806 4M ADCSI10us Average100 SI30000us medium.xlsx
#130806 4M ADCSI10us Average200 SI1000us medium.xlsx

wb = load_workbook(filename)
sheet = wb.get_sheet_by_name(name = 'Tabelle1')

def FindRows():   
    dimension=sheet.calculate_dimension()
    position=dimension.find(':')
    rows=int(dimension[position+2:])
    return rows

x = range(1)
y = range(1)
for i in range(2,FindRows()):
     x.append(sheet.cell(row = i, column = 4).value)
     y.append(sheet.cell(row = i, column = 5).value)

#plot the results     
figure(1)
plot(x,y,'ro')
title(filename)
xlabel('Angle [degree]')
ylabel('Amplitude [mm]')
grid(True)

figure(2)
plot(x,'ro')
title(filename)
xlabel('Sample number []')
ylabel('Angle [degree]')
grid(True)

figure(3)
plot(y,'ro')
title(filename)
xlabel('Sample number []')
ylabel('Amplitude [mm]')
grid(True)